# -*- coding: utf-8 -*-
"""Suma los informes de varias partidas en un solo marcador.

Sirve para Wayground (antes Quizizz) y para Kahoot: las dos puntuan cada
partida por separado. Si la sesion se juega en cuatro
partidas, hacen falta cuatro PIN y salen cuatro marcadores sueltos. Esto los
cruza por apodo y saca uno solo.

    py tools/concurso-marcador.py <carpeta-con-los-informes>

La carpeta lleva los .xlsx que la plataforma deja al descargar el informe de
cada partida. El orden no importa. Se puede usar con una sesion o con el curso
entero: una carpeta por sesion, o todas juntas para el acumulado.

Sobre los apodos: se normalizan (sin mayusculas, sin tildes, sin espacios) para
que «Ana P.» y «ana p» cuenten como la misma persona. Aun asi, quien entre con
un apodo distinto en cada partida sale como personas distintas, y eso no lo
puede arreglar ningun programa: hay que decirlo en voz alta antes de empezar.

Las dos plataformas han cambiado el formato de sus informes varias veces, y
el de Wayground no lo he podido comprobar contra un export real. Por eso, en
vez de dar por buena una plantilla, el guion busca en cada hoja una que tenga columna
de jugador y columna de puntos. Si no la encuentra, lo dice y sigue con las
demas en lugar de fallar entero.
"""
import os
import sys
import unicodedata

sys.stdout.reconfigure(encoding='utf-8')

JUGADOR = ('player', 'jugador', 'nickname', 'apodo', 'name', 'nombre',
           'student', 'estudiante', 'participant', 'participante')
PUNTOS = ('total score', 'score', 'puntuacion', 'puntos', 'puntaje',
          'accuracy', 'final score')
ACIERTOS = ('correct answers', 'correct', 'respuestas correctas', 'correctas',
            'questions correct')


def limpia(x):
    if x is None:
        return ''
    s = unicodedata.normalize('NFKD', str(x))
    return ''.join(c for c in s if not unicodedata.combining(c)).strip().lower()


def clave(apodo):
    """El apodo, reducido a lo que sobrevive a escribirlo mal."""
    s = limpia(apodo)
    return ''.join(c for c in s if c.isalnum())


def busca_columna(cabecera, candidatos):
    for i, c in enumerate(cabecera):
        t = limpia(c)
        for cand in candidatos:
            if t == cand or t.startswith(cand):
                return i
    return None


def lee_partida(ruta):
    """Devuelve {clave: (apodo, puntos, aciertos)} o None si no hay marcador."""
    from openpyxl import load_workbook
    wb = load_workbook(ruta, data_only=True)
    for hoja in wb.worksheets:
        filas = list(hoja.iter_rows(values_only=True))
        for n, fila in enumerate(filas[:12]):          # la cabecera va arriba
            ij = busca_columna(fila, JUGADOR)
            ip = busca_columna(fila, PUNTOS)
            if ij is None or ip is None:
                continue
            ia = busca_columna(fila, ACIERTOS)
            gente = {}
            for f in filas[n + 1:]:
                if ij >= len(f) or not f[ij]:
                    continue
                apodo = str(f[ij]).strip()
                if not clave(apodo):
                    continue
                try:
                    pts = float(f[ip]) if ip < len(f) and f[ip] is not None else 0.0
                except (TypeError, ValueError):
                    continue
                try:
                    ac = int(f[ia]) if ia is not None and ia < len(f) and f[ia] is not None else 0
                except (TypeError, ValueError):
                    ac = 0
                gente[clave(apodo)] = (apodo, pts, ac)
            if gente:
                return hoja.title, gente
    return None


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__.strip().splitlines()[2].strip() or 'Falta la carpeta.')
    carpeta = sys.argv[1]
    if not os.path.isdir(carpeta):
        sys.exit('No existe la carpeta: ' + carpeta)
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        sys.exit('Falta openpyxl:  py -m pip install openpyxl')

    archivos = sorted(f for f in os.listdir(carpeta) if f.lower().endswith('.xlsx')
                      and not f.startswith('~$'))
    if not archivos:
        sys.exit('No hay ningun .xlsx en ' + carpeta)

    partidas, sin_leer = [], []
    for f in archivos:
        r = lee_partida(os.path.join(carpeta, f))
        if r is None:
            sin_leer.append(f)
        else:
            partidas.append((f, r[0], r[1]))

    if not partidas:
        print('Ninguno de los %d archivos tiene un marcador reconocible.' % len(archivos))
        print('Abre uno y mira que columnas trae: hacen falta uno de %s y uno de %s.'
              % (' / '.join(JUGADOR[:3]), ' / '.join(PUNTOS[:3])))
        sys.exit(1)

    print('PARTIDAS LEIDAS')
    for f, hoja, gente in partidas:
        print('  %-46s %2d personas  (hoja «%s»)' % (f[:46], len(gente), hoja))
    for f in sin_leer:
        print('  %-46s  sin marcador reconocible' % f[:46])

    total = {}
    for _, _, gente in partidas:
        for k, (apodo, pts, ac) in gente.items():
            if k not in total:
                total[k] = {'apodo': apodo, 'pts': 0.0, 'ac': 0, 'jugadas': 0}
            total[k]['pts'] += pts
            total[k]['ac'] += ac
            total[k]['jugadas'] += 1
            if len(apodo) > len(total[k]['apodo']):
                total[k]['apodo'] = apodo

    orden = sorted(total.values(), key=lambda d: (-d['pts'], -d['ac'], d['apodo'].lower()))
    n = len(partidas)

    print('\nMARCADOR ACUMULADO  ·  %d partidas  ·  %d personas' % (n, len(orden)))
    print('  #   apodo                        puntos  aciertos  partidas')
    for i, d in enumerate(orden, 1):
        aviso = '' if d['jugadas'] == n else '   <- le faltan %d' % (n - d['jugadas'])
        print('  %-3d %-28s %7d %9d %6d/%d%s'
              % (i, d['apodo'][:28], round(d['pts']), d['ac'], d['jugadas'], n, aviso))

    faltan = [d for d in orden if d['jugadas'] < n]
    if faltan:
        cuantas = ('1 persona no jugó' if len(faltan) == 1
                   else '%d personas no jugaron' % len(faltan))
        print('\n%s las %d partidas. Puede ser que se cayera la'
              % (cuantas, n))
        print('conexion, o que entraran con otro apodo: eso las duplica en la lista.')


if __name__ == '__main__':
    main()
